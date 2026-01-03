"""
Excel Exporter for SCDO
Exports syllabi and CO-PO-PSO mappings to Excel format for accreditation
"""

from typing import Dict, Any, List, Optional
import logging
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logging.warning("openpyxl not installed. Install with: pip install openpyxl")


class ExcelExporter:
    """Export syllabus and CO-PO-PSO mapping to Excel format"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl required for Excel export. Install with: pip install openpyxl")
            
    def export_complete_syllabus(
        self,
        syllabus_data: Dict[str, Any],
        output_path: str,
        include_mapping: bool = True,
        include_rubrics: bool = True
    ) -> bool:
        """
        Export complete syllabus to Excel with multiple sheets
        
        Args:
            syllabus_data: Syllabus structure
            output_path: Output Excel file path
            include_mapping: Include CO-PO-PSO mapping sheet
            include_rubrics: Include rubrics sheet
            
        Returns:
            True if successful
        """
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create sheets
            self._create_overview_sheet(wb, syllabus_data)
            self._create_units_sheet(wb, syllabus_data)
            self._create_outcomes_sheet(wb, syllabus_data)
            
            if include_mapping and 'co_po_mapping' in syllabus_data:
                self._create_mapping_sheet(wb, syllabus_data)
                
            if include_rubrics and 'rubrics' in syllabus_data:
                self._create_rubrics_sheet(wb, syllabus_data)
                
            self._create_assessment_sheet(wb, syllabus_data)
            self._create_references_sheet(wb, syllabus_data)
            
            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Excel exported successfully to {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Excel export failed: {e}")
            return False
            
    def export_mapping_only(
        self,
        syllabus_data: Dict[str, Any],
        output_path: str,
        include_pso: bool = True
    ) -> bool:
        """
        Export only CO-PO-PSO mapping matrix to Excel
        
        Args:
            syllabus_data: Syllabus with mapping data
            output_path: Output Excel file path
            include_pso: Include PSO mapping
            
        Returns:
            True if successful
        """
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "CO-PO-PSO Mapping"
            
            # Create detailed mapping matrix
            self._create_detailed_mapping(ws, syllabus_data, include_pso)
            
            wb.save(output_path)
            self.logger.info(f"Mapping exported to {output_path}")
            return True
            
        except Exception as e:
            self.logger.error(f"Mapping export failed: {e}")
            return False
            
    def _create_overview_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create course overview sheet"""
        ws = wb.create_sheet("Course Overview")
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=14)
        
        # Course details
        ws['A1'] = "Course Information"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws.merge_cells('A1:B1')
        
        # Add course details
        details = [
            ("Course Title", data.get('course_title', 'N/A')),
            ("Course Code", data.get('course_code', 'N/A')),
            ("Credits", data.get('credits', 'N/A')),
            ("Department", data.get('department', 'N/A')),
            ("Level", data.get('level', 'Undergraduate')),
        ]
        
        row = 2
        for label, value in details:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
        # Course overview
        row += 1
        ws[f'A{row}'] = "Course Overview"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        overview = data.get('overview', 'Not provided')
        ws[f'A{row}'] = overview
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 60
        
    def _create_units_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create units/modules sheet"""
        ws = wb.create_sheet("Units & Topics")
        
        # Headers
        headers = ["Unit No.", "Unit Title", "Topics", "Hours"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            
        # Add units
        units = data.get('units', [])
        row = 2
        for unit in units:
            ws.cell(row=row, column=1, value=unit.get('unit_number', ''))
            ws.cell(row=row, column=2, value=unit.get('title', ''))
            
            # Topics as bulleted list
            topics = unit.get('topics', [])
            topics_text = '\n'.join(f"• {topic}" for topic in topics)
            topics_cell = ws.cell(row=row, column=3, value=topics_text)
            topics_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            ws.cell(row=row, column=4, value=unit.get('hours', ''))
            row += 1
            
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 10
        
    def _create_outcomes_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create learning outcomes sheet"""
        ws = wb.create_sheet("Learning Outcomes")
        
        # Headers
        headers = ["CO Code", "Description", "Bloom's Level"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            
        # Add outcomes
        outcomes = data.get('learning_outcomes', [])
        row = 2
        for outcome in outcomes:
            if isinstance(outcome, dict):
                ws.cell(row=row, column=1, value=outcome.get('code', ''))
                desc_cell = ws.cell(row=row, column=2, value=outcome.get('description', ''))
                desc_cell.alignment = Alignment(wrap_text=True)
                ws.cell(row=row, column=3, value=outcome.get('bloom_level', ''))
            else:
                ws.cell(row=row, column=1, value=f"CO{row-1}")
                desc_cell = ws.cell(row=row, column=2, value=str(outcome))
                desc_cell.alignment = Alignment(wrap_text=True)
                ws.cell(row=row, column=3, value='')
            row += 1
            
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 15
        
    def _create_mapping_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create CO-PO-PSO mapping matrix sheet"""
        ws = wb.create_sheet("CO-PO-PSO Mapping")
        self._create_detailed_mapping(ws, data, include_pso=True)
        
    def _create_detailed_mapping(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        data: Dict[str, Any],
        include_pso: bool = True
    ):
        """Create detailed mapping matrix with formatting"""
        
        # Get mapping data
        co_po_mapping = data.get('co_po_mapping', {})
        if not co_po_mapping:
            ws['A1'] = "No mapping data available"
            return
            
        # Get all POs and PSOs
        all_pos = set()
        for po_dict in co_po_mapping.values():
            all_pos.update(po_dict.keys())
        all_pos = sorted([po for po in all_pos if po.startswith('PO')])
        all_psos = sorted([pso for pso in all_pos if pso.startswith('PSO')])
        
        # Create header
        ws['A1'] = "CO"
        ws['A1'].font = Font(bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center')
        
        col = 2
        # PO headers
        for po in all_pos:
            if not po.startswith('PSO'):
                cell = ws.cell(row=1, column=col)
                cell.value = po
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                col += 1
                
        # PSO headers if included
        if include_pso and all_psos:
            for pso in all_psos:
                cell = ws.cell(row=1, column=col)
                cell.value = pso
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
                col += 1
                
        # Add mapping data
        row = 2
        for co, po_dict in sorted(co_po_mapping.items()):
            ws.cell(row=row, column=1, value=co).font = Font(bold=True)
            
            col = 2
            for po in all_pos:
                if not po.startswith('PSO'):
                    value = po_dict.get(po, 0)
                    cell = ws.cell(row=row, column=col)
                    
                    if value > 0:
                        cell.value = value
                        # Color code by correlation level
                        if value == 3:
                            cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                        elif value == 2:
                            cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                        elif value == 1:
                            cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                    else:
                        cell.value = "-"
                        
                    cell.alignment = Alignment(horizontal='center')
                    col += 1
                    
            # PSO mapping
            if include_pso and all_psos:
                for pso in all_psos:
                    value = po_dict.get(pso, 0)
                    cell = ws.cell(row=row, column=col)
                    cell.value = value if value > 0 else "-"
                    cell.alignment = Alignment(horizontal='center')
                    if value > 0:
                        cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                    col += 1
                    
            row += 1
            
        # Add legend
        row += 2
        ws.cell(row=row, column=1, value="Legend:").font = Font(bold=True)
        row += 1
        legend_data = [
            ("3", "High Correlation", "90EE90"),
            ("2", "Medium Correlation", "FFFF99"),
            ("1", "Low Correlation", "FFE4B5"),
        ]
        for val, desc, color in legend_data:
            ws.cell(row=row, column=1, value=val)
            ws.cell(row=row, column=2, value=desc)
            ws.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            row += 1
            
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 8
        for col in range(2, col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 6
            
    def _create_rubrics_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create assessment rubrics sheet"""
        ws = wb.create_sheet("Assessment Rubrics")
        
        rubrics = data.get('rubrics', {})
        if not rubrics:
            ws['A1'] = "No rubrics available"
            return
            
        row = 1
        for component_name, rubric in rubrics.items():
            # Component header
            ws.cell(row=row, column=1, value=component_name.replace('_', ' ').title())
            ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
            ws.cell(row=row, column=1).fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            # Component details
            ws.cell(row=row, column=1, value="Type:")
            ws.cell(row=row, column=2, value=rubric.get('type', 'N/A'))
            ws.cell(row=row, column=3, value="Total Marks:")
            ws.cell(row=row, column=4, value=rubric.get('total_marks', 'N/A'))
            row += 1
            
            # Criteria headers
            headers = ["Criterion", "Weightage (%)", "Level", "Description"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            row += 1
            
            # Criteria details
            for criterion in rubric.get('criteria', []):
                criterion_name = criterion.get('name', '')
                weightage = criterion.get('weightage', 0)
                levels = criterion.get('levels', {})
                
                first_level = True
                for level, description in levels.items():
                    ws.cell(row=row, column=1, value=criterion_name if first_level else "")
                    ws.cell(row=row, column=2, value=weightage if first_level else "")
                    ws.cell(row=row, column=3, value=level)
                    desc_cell = ws.cell(row=row, column=4, value=description)
                    desc_cell.alignment = Alignment(wrap_text=True)
                    first_level = False
                    row += 1
                    
            row += 2  # Space between components
            
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        
    def _create_assessment_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create assessment pattern sheet"""
        ws = wb.create_sheet("Assessment Pattern")
        
        assessment = data.get('assessment_pattern', {})
        if not assessment:
            ws['A1'] = "No assessment data available"
            return
            
        # Headers
        ws['A1'] = "Assessment Component"
        ws['B1'] = "Weightage (%)"
        for cell in [ws['A1'], ws['B1']]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
        row = 2
        # Internal assessment
        internal = assessment.get('internal', {})
        ws.cell(row=row, column=1, value="Internal Assessment").font = Font(bold=True)
        ws.cell(row=row, column=2, value=internal.get('weightage', 0))
        row += 1
        
        for component, marks in internal.get('components', {}).items():
            ws.cell(row=row, column=1, value=f"  • {component.replace('_', ' ').title()}")
            ws.cell(row=row, column=2, value=marks)
            row += 1
            
        # External assessment
        row += 1
        external = assessment.get('external', {})
        ws.cell(row=row, column=1, value="External Assessment").font = Font(bold=True)
        ws.cell(row=row, column=2, value=external.get('weightage', 0))
        row += 1
        
        for component, marks in external.get('components', {}).items():
            ws.cell(row=row, column=1, value=f"  • {component.replace('_', ' ').title()}")
            ws.cell(row=row, column=2, value=marks)
            row += 1
            
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        
    def _create_references_sheet(self, wb: openpyxl.Workbook, data: Dict[str, Any]):
        """Create references sheet"""
        ws = wb.create_sheet("References")
        
        # Header
        ws['A1'] = "References & Resources"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        ws.merge_cells('A1:B1')
        
        references = data.get('references', [])
        row = 2
        for i, ref in enumerate(references, 1):
            ws.cell(row=row, column=1, value=f"{i}.")
            ref_cell = ws.cell(row=row, column=2, value=ref)
            ref_cell.alignment = Alignment(wrap_text=True)
            row += 1
            
        # Auto-adjust columns
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 80
